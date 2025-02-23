import matplotlib.pyplot as plt
import os
import pandas as pd
import numpy as np
from scipy.signal import argrelextrema
from scipy.fftpack import fft, fftfreq
import scipy.signal
from scipy.signal import find_peaks
from openpyxl import load_workbook
import openpyxl

def find_stable_regions(data, window_size=5, threshold=0.3):
    """
    找出數據中的平穩區域
    window_size: 滑動窗口大小
    threshold: 判定為平穩的變化閾值
    """
    variations = []
    means = []
    
    # 計算每個窗口的變化程度
    for i in range(len(data) - window_size + 1):
        window = data[i:i+window_size]
        variation = np.std(window)  # 使用標準差來衡量變化程度
        mean = np.mean(window)
        variations.append(variation)
        means.append(mean)
    
    # 找出變化小於閾值的區域
    stable_regions = []
    current_region = []
    
    for i, var in enumerate(variations):
        if var < threshold:
            current_region.append(means[i])
        elif current_region:
            stable_regions.append(np.mean(current_region))
            current_region = []
    
    if current_region:
        stable_regions.append(np.mean(current_region))
    
    print("平穩區域: ", stable_regions)
    return np.array(stable_regions)

def filter_stable_regions(time, elec, begin_angle, rest_angle, begin_tolerance=20, rest_tolerance=4):
    """
    只過濾序列開頭和結尾的平穩區段
    
    Parameters:
    - time: 時間序列
    - elec: 角度序列
    - begin_angle: 起始角度
    - rest_angle: 靜息角度
    - tolerance: 判定為平穩區域的閾值
    
    Returns:
    - filtered_time: 過濾後的時間序列
    - filtered_elec: 過濾後的角度序列
    """
    # 找出序列開頭的平穩區段結束點
    start_idx = 0
    for i in range(len(elec)):
        if abs(elec[i] - begin_angle) > begin_tolerance:
            start_idx = i
            break
    
    # 從後往前找出序列結尾的平穩區段起始點
    end_idx = len(elec) - 1
    for i in range(len(elec)-1, -1, -1):
        if abs(elec[i] - rest_angle) > rest_tolerance:
            end_idx = i
            break
    
    # 過濾數據
    filtered_time = time[start_idx:end_idx+1]
    filtered_elec = elec[start_idx:end_idx+1]
    
    return filtered_time, filtered_elec

def analyze_folder(folder_path):
    """
    分析特定資料夾中的數據並返回結果
    """
    # 讀取數據
    file_elec = r'elec.txt'
    dir_elec = os.path.join(folder_path, file_elec)
    
    df_elec = pd.read_csv(dir_elec, sep=",")
    time = np.array(df_elec['Time'])
    if 'L' in dir_elec:
        elec = np.array(df_elec['Linear Transformer Gonio G'])
    else:
        elec = np.array(180 - df_elec['Linear Transformer Gonio G'])
    
    # 調整角度值
    adjustment = 180 - elec[0]
    elec = elec + adjustment

    # 數據平滑處理
    elec = scipy.signal.savgol_filter(elec, 21, 4, mode='nearest')

    # Reset time to start from zero
    time_zeroed = time - time[0]

    # 找出局部極值
    max_index = argrelextrema(elec, np.greater, order=10)
    min_index = argrelextrema(elec, np.less, order=10)

    fig = plt.figure(figsize=(8, 6))
    plt.plot(time_zeroed, elec)
    plt.scatter(
        time_zeroed[max_index],
        elec[max_index],
        c='red',
        label='max',
        s=25
    )

    plt.scatter(
        time_zeroed[min_index],
        elec[min_index],
        c='blue',
        label='min',
        s=10
    )

    plt.legend(loc=1, fontsize='large')
    plt.title('Angle Variation Over Times')
    plt.xlabel('Time (zeroed)')
    plt.ylabel('Angles')
    plt.show()

    # 找出平穩區域
    stable_regions = find_stable_regions(elec[min_index])
    stable_regions = np.sort(stable_regions)

    # 確定起始角度和靜息角度
    if len(stable_regions) >= 2:
        rest_angle = stable_regions[0]
        begin_angle = stable_regions[-1]
    elif len(stable_regions) == 1:
        if stable_regions[0] > 165:
            begin_angle = stable_regions[0]
            rest_angle = elec[-1]
        else:
            rest_angle = stable_regions[0]
            begin_angle = max(elec)
    else:
        begin_angle = max(elec)
        rest_angle = elec[-1]

    print(f"重新處理後的角度: (起始角度){begin_angle}, (靜息角度){rest_angle}")

    # 過濾平穩區域
    filtered_time, filtered_elec = filter_stable_regions(time_zeroed, elec, begin_angle, rest_angle)

    # 在過濾後的數據上找極值點
    max_index_filtered = argrelextrema(filtered_elec, np.greater, order=50)[0]
    min_index_filtered = argrelextrema(filtered_elec, np.less, order=50)[0]

    # 計算相關參數
    threshold = rest_angle
    a = filtered_elec[max_index_filtered[0]]
    b = filtered_elec[min_index_filtered[0]]
    c = filtered_elec[min_index_filtered[1]]

    a0 = begin_angle - threshold
    a1 = begin_angle - b
    a2 = a - b
    a3 = a - threshold
    a4 = a - c

    num_waves = len(max_index_filtered)

    p1 = a1 / (1.6 * a0)
    p4 = a3
    p5 = a4 / (1.6 * a3)

    # 繪圖
    plt.figure(figsize=(12, 8))
    plt.plot(time_zeroed, elec, 'black', label='Original Signal')
    plt.plot(filtered_time, filtered_elec, 'black', label='Filtered Signal', alpha=0.2)


    plt.scatter(filtered_time[max_index_filtered], filtered_elec[max_index_filtered], 
               c='red', s=30, label='Local Max')
    plt.scatter(filtered_time[min_index_filtered], filtered_elec[min_index_filtered], 
               c='blue', s=30, label='Local Min')

    # 添加平穩區域的標記
    plt.axhline(y=rest_angle, color='#F5D5B1', linestyle='--', label='Rest Angle')
    plt.axhline(y=begin_angle, color='#F4C2B1', linestyle='--', label='Begin Angle')

    # 添加關鍵點
    plt.axhline(y=a, color='#BCF5D4', linestyle='-.', label='a')
    plt.axhline(y=b, color='#A5D7CE', linestyle='-.', label='b')
    plt.axhline(y=c, color='#96BBD6', linestyle='-.', label='c')

    # 添加計算結果
    bbox_props = dict(boxstyle="round,pad=0.3", edgecolor="none", facecolor="white", alpha=0.8)
    result_text = (f"A0={a0:.3f}\nA1={a1:.3f}\nA2={a2:.3f}\nA3={a3:.3f}\nA4={a4:.3f}")
    plt.text(0.05, 0.22, result_text, transform=plt.gca().transAxes, fontsize=12, 
             verticalalignment='top', bbox=bbox_props)

    result_text = (f"p1={p1:.3f}\np4={p4:.3f}\np5={p5:.3f}")
    plt.text(0.27, 0.22, result_text, transform=plt.gca().transAxes, fontsize=12, 
             verticalalignment='top', bbox=bbox_props)

    plt.legend(loc='upper right', fontsize='medium')
    plt.title('Angle Variation Over Time')
    plt.xlabel('Time (s)')
    plt.ylabel('Angle (degrees)')
    plt.ylim(0, 190)
    plt.grid(True, alpha=0.3)

    # 保存圖片
    output_path = os.path.join(folder_path, 'para.png')
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.show()

    # 收集結果
    results = {
        '起始角度': begin_angle,
        '靜息角度': rest_angle,
        'a': a,
        'b': b,
        'c': c,
        'A0': a0,
        'A1': a1,
        'A2': a2,
        'A3': a3,
        'A4': a4,
        'Number of waves': num_waves,
        'p1': p1,
        'p4': p4,
        'p5': p5
    }
    
    return results

def export_to_excel(folder_name, data_dict, excel_path='para_fix.xlsx'):
    """
    將分析結果輸出到Excel檔案
    """
    # 定義列的順序
    row_order = [
        '起始角度', '靜息角度', 'a', 'b', 'c',
        'A0', 'A1', 'A2', 'A3', 'A4',
        'Number of waves', 'p1', 'p4', 'p5'
    ]
    
    # 檢查檔案是否存在
    if not os.path.exists(excel_path):
        # 創建新的工作簿並添加表頭
        df = pd.DataFrame(columns=[''] + [folder_name])
        df.loc[:, ''] = row_order
        df.to_excel(excel_path, index=False)
    
    # 載入現有的工作簿
    book = load_workbook(excel_path)
    sheet = book.active
    
    # 找到下一個空白欄
    next_col = sheet.max_column + 1
    
    # 寫入資料夾名稱
    sheet.cell(row=1, column=next_col, value=folder_name)
    
    # 寫入數據
    for i, row_name in enumerate(row_order, start=2):
        value = data_dict.get(row_name, '')
        sheet.cell(row=i, column=next_col, value=value)
    
    # 儲存工作簿
    book.save(excel_path)


def get_available_paths(base_folder):
    """
    獲取所有可用的資料路徑
    """
    available_paths = []
    for folder_name in os.listdir(base_folder):
        folder_path = os.path.join(base_folder, folder_name)
        for fname in os.listdir(folder_path):
            fname_path = os.path.join(folder_path, fname)
            if os.path.isdir(fname_path) and os.path.exists(os.path.join(fname_path, 'elec.txt')):
                available_paths.append((fname_path, fname))
    return available_paths

def process_selected_folder(path_info):
    """
    處理選定的資料夾
    """
    fname_path, fname = path_info
    try:
        results = analyze_folder(fname_path)
        export_to_excel(fname, results)
        print(f"成功處理 {fname}")
    except Exception as e:
        print(f"處理 {fname} 時發生錯誤: {str(e)}")

def main():
    # 設定基礎資料夾路徑
    base_folder = r"D:/para/001"  # 請根據實際路徑調整
    print(f"base_folder: {base_folder}")

    # 獲取所有可用的路徑
    available_paths = get_available_paths(base_folder)

    # 列出所有可用的路徑
    print("\n可用的資料路徑：")
    for i, (path, fname) in enumerate(available_paths, 1):
        print(f"{i}. {path} ({fname})")

    # 讓使用者選擇要處理的路徑
    while True:
        try:
            choice = input("\n請選擇要處理的路徑編號 (輸入 'q' 退出): ")
            
            if choice.lower() == 'q':
                print("程式結束")
                break

            choice = int(choice)
            if 1 <= choice <= len(available_paths):
                selected_path = available_paths[choice - 1]
                print(f"\n處理路徑: {selected_path[0]}")
                process_selected_folder(selected_path)
            else:
                print("無效的選擇，請輸入有效的編號")
        except ValueError:
            print("請輸入有效的數字或 'q' 退出")

if __name__ == "__main__":
    main()