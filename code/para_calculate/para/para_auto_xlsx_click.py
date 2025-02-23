import matplotlib.pyplot as plt
import os
import pandas as pd
import numpy as np
from scipy.signal import argrelextrema
import scipy.signal
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox

class PointSelector:
    def __init__(self, time, signal):
        self.time = time
        self.signal = signal
        self.points = {'begin': None, 'a': None, 'b': None, 'c': None}
        self.current_point = None
        self.fig = None
        self.ax = None
        
    def onclick(self, event):
        if event.inaxes != self.ax:
            return
        
        if self.current_point is None:
            return
            
        x, y = event.xdata, event.ydata
        self.points[self.current_point] = (x, y)
        
        # Update plot
        self.ax.clear()
        self.plot_data()
        plt.draw()
        
        # Move to next point
        if self.current_point == 'begin':
            self.current_point = 'a'
            plt.title('Click to select point a (first peak)')
        elif self.current_point == 'a':
            self.current_point = 'b'
            plt.title('Click to select point b (first valley)')
        elif self.current_point == 'b':
            self.current_point = 'c'
            plt.title('Click to select point c (second valley)')
        else:
            plt.close()
            
    def plot_data(self):
        self.ax.plot(self.time, self.signal, 'black')
        for point_name, point in self.points.items():
            if point is not None:
                color = 'red' if point_name in ['begin', 'a'] else 'blue'
                self.ax.scatter(point[0], point[1], c=color, s=100)
                self.ax.annotate(point_name, (point[0], point[1]))
        self.ax.grid(True, alpha=0.3)
        self.ax.set_xlabel('Time (s)')
        self.ax.set_ylabel('Angle (degrees)')
        # self.ax.set_ylim(0, 190)
        
    def select_points(self):
        self.fig, self.ax = plt.subplots(figsize=(12, 8))
        self.current_point = 'begin'
        self.plot_data()
        plt.title('Click to select beginning point')
        self.fig.canvas.mpl_connect('button_press_event', self.onclick)
        plt.show()
        return self.points

def find_stable_regions(data, window_size=5, threshold=0.3):
    """
    找出數據中的平穩區域
    window_size: 滑動窗口大小
    threshold: 判定為平穩的變化閾值
    """
    variations = []
    means = []
    
    # 計算每個窗口的變化程度
    for i in range(len(data) - window_size + 1):
        window = data[i:i+window_size]
        variation = np.std(window)  # 使用標準差來衡量變化程度
        mean = np.mean(window)
        variations.append(variation)
        means.append(mean)
    
    # 找出變化小於閾值的區域
    stable_regions = []
    current_region = []
    
    for i, var in enumerate(variations):
        if var < threshold:
            current_region.append(means[i])
        elif current_region:
            stable_regions.append(np.mean(current_region))
            current_region = []
    
    if current_region:
        stable_regions.append(np.mean(current_region))
    
    print("平穩區域: ", stable_regions)
    return np.array(stable_regions)

def filter_stable_regions(time, elec, begin_angle, rest_angle, begin_tolerance=15, rest_tolerance=4):
    """
    只過濾序列開頭和結尾的平穩區段
    
    Parameters:
    - time: 時間序列
    - elec: 角度序列
    - begin_angle: 起始角度
    - rest_angle: 靜息角度
    - tolerance: 判定為平穩區域的閾值
    
    Returns:
    - filtered_time: 過濾後的時間序列
    - filtered_elec: 過濾後的角度序列
    """
    # 找出序列開頭的平穩區段結束點
    start_idx = 0
    for i in range(len(elec)):
        if abs(elec[i] - begin_angle) > begin_tolerance:
            start_idx = i
            break
    
    # 從後往前找出序列結尾的平穩區段起始點
    end_idx = len(elec) - 1
    for i in range(len(elec)-1, -1, -1):
        if abs(elec[i] - rest_angle) > rest_tolerance:
            end_idx = i
            break
    
    # 過濾數據
    filtered_time = time[start_idx:end_idx+1]
    filtered_elec = elec[start_idx:end_idx+1]
    
    return filtered_time, filtered_elec

def auto_analyze(time, signal):
    """自動分析信號"""
    # 找出局部極值
    max_index = argrelextrema(signal, np.greater, order=50)[0]
    min_index = argrelextrema(signal, np.less, order=50)[0]
    
    # 找出平穩區域
    stable_regions = find_stable_regions(signal[min_index])
    stable_regions = np.sort(stable_regions)
    
    # 確定起始角度和靜息角度
    if len(stable_regions) >= 2:
        rest_angle = stable_regions[0]
        begin_angle = stable_regions[-1]
    elif len(stable_regions) == 1:
        if stable_regions[0] > 165:
            begin_angle = stable_regions[0]
            rest_angle = signal[-1]
        else:
            rest_angle = stable_regions[0]
            begin_angle = signal[0]
    else:
        # begin_angle = max(signal)
        begin_angle = signal[0]
        rest_angle = signal[-1]

    # 過濾平穩區域
    filtered_time, filtered_signal = filter_stable_regions(time, signal, begin_angle, rest_angle)
    
    # 在過濾後的數據上找極值點
    max_index_filtered = argrelextrema(filtered_signal, np.greater, order=50)[0]
    min_index_filtered = argrelextrema(filtered_signal, np.less, order=50)[0]
    
    # 檢查是否找到足夠的極值點
    if len(max_index_filtered) == 0 or len(min_index_filtered) == 0:
        print(f"警告：未找到足夠的波峰波谷。轉換為手動模式。")
        return None
        
    # 找到對應的時間點
    a_time = filtered_time[max_index_filtered[0]]
    b_time = filtered_time[min_index_filtered[0]]
    
    # 如果找到第二個波谷，使用它；否則使用最後一個數據點
    if len(min_index_filtered) > 1:
        c_time = filtered_time[min_index_filtered[1]]
        c_value = filtered_signal[min_index_filtered[1]]
    else:
        c_time = filtered_time[-1]
        c_value = filtered_signal[-1]
        print("警告：只找到一個波谷，使用最後一個數據點作為 c 點")
    
    return {
        'begin_angle': begin_angle,
        'rest_angle': rest_angle,
        'a': filtered_signal[max_index_filtered[0]],
        'b': filtered_signal[min_index_filtered[0]],
        'c': c_value,
        'a_time': a_time,
        'b_time': b_time,
        'c_time': c_time,
        'filtered_time': filtered_time,
        'filtered_signal': filtered_signal,
        'max_index': max_index_filtered,
        'min_index': min_index_filtered
    }

def analyze_folder(folder_path, manual_mode=False):
    print("分析資料夾：", folder_path)
    """分析特定資料夾中的數據並返回結果"""
    # 讀取數據
    file_hpe = r'ae_half.xlsx'
    dir_hpe = os.path.join(folder_path, file_hpe)
    df_hpe = pd.read_excel(dir_hpe)
    hpe = np.array(df_hpe['Processed_Output'])
    time = np.arange(len(hpe))
    
    # # 調整角度值
    # adjustment = 180 - hpe[0]
    # hpe = hpe + adjustment
    
    # 數據平滑處理
    hpe = scipy.signal.savgol_filter(hpe, 21, 3, mode='nearest')
    time_zeroed = time - time[0]
    
    if not manual_mode:
        # 自動分析模式
        results = auto_analyze(time_zeroed, hpe)
        if results is None:
            print("自動分析失敗，切換到手動模式")
            return analyze_folder(folder_path, manual_mode=True)
    else:
        # 手動選點模式
        results = manual_analyze(time_zeroed, hpe)
    
    # 繪製結果圖並取得更新後的結果
    results = plot_results(time_zeroed, hpe, results)
    
    # 詢問是否保存結果
    root = tk.Tk()
    root.withdraw()  # 隱藏主窗口
    save = messagebox.askyesno("確認", "是否要保存這個分析結果？")
    
    if not save:
        plt.close()
        if not manual_mode:
            # 如果不保存自動分析結果，切換到手動模式
            return analyze_folder(folder_path, manual_mode=True)
        return None
        
    plt.close()
    return results

def manual_analyze(time, signal):
    """手動選點分析信號"""
    selector = PointSelector(time, signal)
    points = selector.select_points()
    
    begin_angle = points['begin'][1]
    rest_angle = signal[-1]  # 使用序列末端作為靜息角度
    
    return {
        'begin_angle': begin_angle,
        'rest_angle': rest_angle,
        'a': points['a'][1],
        'b': points['b'][1],
        'c': points['c'][1],
        'a_time': points['a'][0],
        'b_time': points['b'][0],
        'c_time': points['c'][0],
        'filtered_time': time,
        'filtered_signal': signal,
        'max_index': [],
        'min_index': []
    }

def plot_results(time, signal, results):
    """繪製分析結果"""
    if results is None:
        return None
        
    plt.figure(figsize=(12, 8))
    plt.plot(time, signal, 'black', label='Signal')
    
    # 添加關鍵點和線條
    plt.axhline(y=results['rest_angle'], color='#F5D5B1', linestyle='--', label='Rest Angle')
    plt.axhline(y=results['begin_angle'], color='#F4C2B1', linestyle='--', label='Begin Angle')
    plt.axhline(y=results['a'], color='#BCF5D4', linestyle='-.', label='a')
    plt.axhline(y=results['b'], color='#A5D7CE', linestyle='-.', label='b')
    plt.axhline(y=results['c'], color='#96BBD6', linestyle='-.', label='c')
    
    # 添加 a、b、c 點的散點標記
    plt.scatter(results['a_time'], results['a'], color='red', s=30, zorder=5, label='Point a')
    plt.scatter(results['b_time'], results['b'], color='blue', s=30, zorder=5, label='Point b')
    plt.scatter(results['c_time'], results['c'], color='blue', s=30, zorder=5, label='Point c')
    
    # 計算並顯示參數
    threshold = results['rest_angle']
    a0 = float(results['begin_angle'] - threshold)
    a1 = float(results['begin_angle'] - results['b'])
    a2 = float(results['a'] - results['b'])
    a3 = float(results['a'] - threshold)
    a4 = float(results['a'] - results['c'])
    
    p1 = float(a1 / (1.6 * a0))
    p4 = float(a3)
    p5 = float(a4 / (1.6 * a3))
    
    # 更新results字典
    results.update({
        'A0': a0,
        'A1': a1,
        'A2': a2,
        'A3': a3,
        'A4': a4,
        'p1': p1,
        'p4': p4,
        'p5': p5,
        'Number of waves': len(results.get('max_index', [])) or 1
    })
    
    # 添加計算結果文本
    bbox_props = dict(boxstyle="round,pad=0.3", edgecolor="none", facecolor="white", alpha=0.8)
    result_text = (f"A0={a0:.3f}\nA1={a1:.3f}\nA2={a2:.3f}\nA3={a3:.3f}\nA4={a4:.3f}")
    plt.text(0.05, 0.22, result_text, transform=plt.gca().transAxes, fontsize=12, 
             verticalalignment='top', bbox=bbox_props)
    
    result_text = (f"p1={p1:.3f}\np4={p4:.3f}\np5={p5:.3f}")
    plt.text(0.27, 0.22, result_text, transform=plt.gca().transAxes, fontsize=12, 
             verticalalignment='top', bbox=bbox_props)
    
    plt.legend(loc='upper right', fontsize='medium')
    plt.title('Angle Variation Over Time')
    plt.xlabel('Time (s)')
    plt.ylabel('Angle (degrees)')
    # plt.ylim(0, 190)
    plt.grid(True, alpha=0.3)
    plt.show()
    
    return results

def export_to_excel(folder_name, data_dict, excel_path='analysis_results_hpe_v3_fix.xlsx'):
    """
    將分析結果輸出到Excel檔案
    """
    # 定義列的順序
    row_order = [
        '起始角度', '靜息角度', 'a', 'b', 'c',
        'A0', 'A1', 'A2', 'A3', 'A4',
        'Number of waves', 'p1', 'p4', 'p5'
    ]
    
    # 創建新的字典並確保所有值都是數值類型
    processed_dict = {
        '起始角度': float(data_dict['begin_angle']),
        '靜息角度': float(data_dict['rest_angle']),
        'a': float(data_dict['a']),
        'b': float(data_dict['b']),
        'c': float(data_dict['c']),
        'A0': float(data_dict['A0']),
        'A1': float(data_dict['A1']),
        'A2': float(data_dict['A2']),
        'A3': float(data_dict['A3']),
        'A4': float(data_dict['A4']),
        'Number of waves': int(data_dict['Number of waves']),
        'p1': float(data_dict['p1']),
        'p4': float(data_dict['p4']),
        'p5': float(data_dict['p5'])
    }
    
    # 檢查檔案是否存在
    if not os.path.exists(excel_path):
        # 創建新的工作簿並添加表頭
        df = pd.DataFrame(columns=[''] + [folder_name])
        df.loc[:, ''] = row_order
        df.to_excel(excel_path, index=False)
    
    # 載入現有的工作簿
    book = load_workbook(excel_path)
    sheet = book.active
    
    # 找到下一個空白欄
    next_col = sheet.max_column + 1
    
    # 寫入資料夾名稱
    sheet.cell(row=1, column=next_col, value=folder_name)
    
    # 寫入數據
    for i, row_name in enumerate(row_order, start=2):
        value = processed_dict.get(row_name, '')
        sheet.cell(row=i, column=next_col, value=value)
    
    # 儲存工作簿
    book.save(excel_path)

def process_folders(base_folder):
    """處理所有子資料夾並輸出結果到Excel"""
    for folder_name in os.listdir(base_folder):
        folder_path = os.path.join(base_folder, folder_name)
        if os.path.isdir(folder_path):
            for fname in os.listdir(folder_path):
                fname_path = os.path.join(folder_path, fname)
                if os.path.isdir(fname_path) and os.path.exists(os.path.join(fname_path, 'ae_half.xlsx')):
                    try:
                        print(f"正在處理 {fname}")

                        results = analyze_folder(fname_path)
                        if results is not None:  # 只有當使用者確認要保存時才輸出到Excel
                            export_to_excel(fname, results)
                            print(f"成功處理 {fname}")
                        else:
                            print(f"使用者選擇不保存 {fname} 的結果")
                    except Exception as e:
                        print(f"處理 {fname} 時發生錯誤: {str(e)}")

if __name__ == "__main__":
    base_folder = r"D:/para/004"  # 請根據實際路徑調整
    process_folders(base_folder)