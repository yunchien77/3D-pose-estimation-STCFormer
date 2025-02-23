import numpy as np
import pandas as pd
from scipy.signal import find_peaks
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import warnings
warnings.filterwarnings('ignore')

class PeakSelector:
    def __init__(self, data):
        self.data = data
        self.selected_peaks = []
        self.peak_values = []
        self.fig = None
        self.ax = None
        self.finished = False
        
    def onclick(self, event):
        if event.button == 1 and event.inaxes:  # Left click
            x = int(round(event.xdata))
            if x >= 0 and x < len(self.data):
                self.selected_peaks.append(x)
                self.peak_values.append(self.data[x])
                self.ax.plot(x, self.data[x], 'go', markersize=10)
                self.fig.canvas.draw()
        elif event.button == 3:  # Right click to finish
            self.finished = True
            plt.close()
    
    def select_peaks(self):
        self.fig, self.ax = plt.subplots(figsize=(28, 15))
        self.ax.plot(self.data, 'b-')
        self.ax.set_title('Click to select peaks (Right click to finish)')
        self.fig.canvas.mpl_connect('button_press_event', self.onclick)
        plt.show()
        
        return np.array(self.selected_peaks), np.array(self.peak_values)

def read_and_process_data(file_path):
    """
    讀取和處理數據文件，將起始值調整為 180，其餘值根據差值調整
    """
    df = pd.read_csv(file_path, sep=',', names=['Time', 'Linear Transformer Gonio G'])
    angles = pd.to_numeric(df['Linear Transformer Gonio G'], errors='coerce').values
    angles = angles[~np.isnan(angles)]

    if 'R' in file_path:
        angles = 180 - angles

    if len(angles) > 0:
        delta = angles[0]
        angles = 180 + (angles - delta)

    return angles

def filter_decreasing_peaks(peaks, peak_values):
    """
    篩選出符合遞減趨勢的峰值
    """
    if len(peaks) <= 1:
        return peaks, peak_values
    
    filtered_peaks = [peaks[0]]
    filtered_values = [peak_values[0]]
    current_max = peak_values[0]
    
    for i in range(1, len(peaks)):
        if peak_values[i] < current_max:
            filtered_peaks.append(peaks[i])
            filtered_values.append(peak_values[i])
            current_max = peak_values[i]
    
    return np.array(filtered_peaks), np.array(filtered_values)

def find_and_plot_peaks(data, prominence=1, distance=10, height=None, width=None):
    """
    找出並繪製峰值
    """
    peaks, properties = find_peaks(data, 
                               prominence=prominence,
                               distance=distance,
                               height=height,
                               width=width)
    
    original_peaks = peaks.copy()
    original_values = data[peaks].copy()
    peaks, peak_values = filter_decreasing_peaks(peaks, data[peaks])
    
    plt.figure(figsize=(20, 12))
    plt.plot(data, 'b-', label='Original Data', alpha=0.7)
    plt.plot(original_peaks, original_values, 'ro', label='All Detected Peaks', alpha=0.3)
    
    removed_peaks = [p for p in original_peaks if p not in peaks]
    removed_values = [data[p] for p in removed_peaks]
    if removed_peaks:
        plt.plot(removed_peaks, removed_values, 'rx', markersize=10, 
                label='Removed Peaks', alpha=0.7)
    
    plt.plot(peaks, peak_values, 'go', label='Used Peaks', alpha=1)
    
    for i, (peak, value) in enumerate(zip(peaks, peak_values)):
        plt.annotate(f'Peak {i+1}\ny={value:.2f}', 
                    xy=(peak, value),
                    xytext=(10, 10),
                    textcoords='offset points')
    
    plt.title('Data with Detected Peaks (Filtered for Decreasing Trend)')
    plt.xlabel('Time Step')
    plt.ylabel('Angle (degrees)')
    plt.legend()
    plt.grid(True)
    plt.show()
    
    return peaks, peak_values

def exp_func(x, a, b, c):
    """
    指數函數
    """
    return a * np.exp(-b * (x - np.min(x))) + c

def fit_exponential(peaks, peak_values):
    """
    執行指數擬和
    """
    if len(peaks) < 2:
        print("警告：峰值點太少，無法進行可靠的擬和")
        return None, None

    p0 = [
        max(peak_values) - min(peak_values),
        1.0 / (max(peaks) - min(peaks)),
        min(peak_values)
    ]
    
    bounds = (
        [0, 0, min(peak_values) * 0.5],
        [max(peak_values) * 2, 1.0, max(peak_values)]
    )
    
    try:
        popt, pcov = curve_fit(
            exp_func, 
            peaks, 
            peak_values, 
            p0=p0,
            bounds=bounds,
            maxfev=10000
        )
        return popt, pcov
    except Exception as e:
        print(f"擬和錯誤：{str(e)}")
        return None, None

def plot_exponential_fit(data, peaks, peak_values, folder_path):
    """
    繪製指數擬和結果
    """
    popt, pcov = fit_exponential(peaks, peak_values)
    
    if popt is not None:
        plt.figure(figsize=(20, 12))
        plt.plot(data, 'b-', label='Original Data', alpha=0.5)
        plt.plot(peaks, peak_values, 'go', label='Used Peaks', alpha=0.7)
        
        x_fit = np.linspace(min(peaks), max(peaks), 1000)
        y_fit = exp_func(x_fit, *popt)
        plt.plot(x_fit, y_fit, 'r--', label='Exponential Fit')
        
        plt.xlabel('Time Step')
        plt.ylabel('Angle (degrees)')
        plt.title('Exponential Fit to Decreasing Peak Values')
        plt.legend()
        plt.grid(True)
        
        y_min = min(min(data), min(y_fit)) * 0.9
        y_max = max(max(data), max(y_fit)) * 1.1
        plt.ylim(y_min, y_max)
        plt.show()
        
        # Save the plot
        plt.savefig(os.path.join(folder_path, 'exponential_fit.png'))
        plt.close()
        
        return popt
    return None

def export_to_excel(folder_name, data_dict, excel_path='exponential_results.xlsx'):
    """
    將分析結果輸出到Excel檔案
    """
    row_order = [
        'Initial amplitude (a)',
        'Decay rate (b)',
        'Offset (c)',
        'Number of peaks',
        'Peak positions',
        'Peak values'
    ]
    
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=[''] + [folder_name])
        df.loc[:, ''] = row_order
        df.to_excel(excel_path, index=False)
    
    book = load_workbook(excel_path)
    sheet = book.active
    
    next_col = sheet.max_column + 1
    
    sheet.cell(row=1, column=next_col, value=folder_name)
    
    for i, row_name in enumerate(row_order, start=2):
        value = data_dict.get(row_name, '')
        if isinstance(value, (list, np.ndarray)):
            value = ', '.join(map(str, value))
        sheet.cell(row=i, column=next_col, value=value)
    
    book.save(excel_path)

def analyze_file(file_path, folder_path, prominence=1, distance=10, height=None, width=None):
    """
    主要分析函數
    """
    try:
        data = read_and_process_data(file_path)
        
        if len(data) == 0:
            print("警告：沒有有效的數據")
            return None
        
        peaks, peak_values = find_and_plot_peaks(data, 
                                             prominence=prominence,
                                             distance=distance,
                                             height=height,
                                             width=width)
        
        # 顯示確認對話框
        if not messagebox.askyesno("確認", "是否接受目前的峰值檢測結果？"):
            # 如果不接受，進入手動選擇模式
            selector = PeakSelector(data)
            peaks, peak_values = selector.select_peaks()
            if len(peaks) == 0:
                print("未選擇任何峰值點")
                return None
        
        fit_params = plot_exponential_fit(data, peaks, peak_values, folder_path)
        
        if fit_params is not None:
            results = {
                'Initial amplitude (a)': fit_params[0],
                'Decay rate (b)': fit_params[1],
                'Offset (c)': fit_params[2],
                'Number of peaks': len(peaks),
                'Peak positions': peaks,
                'Peak values': peak_values
            }
            return results
        
    except Exception as e:
        print(f"分析過程發生錯誤：{str(e)}")
    return None

def process_folders(base_folder):
    """
    處理所有子資料夾
    """
    root = tk.Tk()
    root.withdraw()  # 隱藏主窗口
    
    for folder_name in os.listdir(base_folder):
        folder_path = os.path.join(base_folder, folder_name)
        print(f"正在處理資料夾: {folder_path}")
        
        for fname in os.listdir(folder_path):
            fname_path = os.path.join(folder_path, fname)
            elec_file = os.path.join(fname_path, 'elec.txt')
            
            if os.path.isdir(fname_path) and os.path.exists(elec_file):
                try:
                    print(f"正在分析: {fname}")
                    results = analyze_file(elec_file, fname_path)
                    
                    if results is not None:
                        if messagebox.askyesno("確認", "是否要將此結果保存到Excel？"):
                            export_to_excel(fname, results)
                            print(f"成功處理並保存 {fname}")
                        else:
                            print(f"已跳過保存 {fname} 的結果")
                    
                except Exception as e:
                    print(f"處理 {fname} 時發生錯誤: {str(e)}")
    
    root.destroy()

def main():
    base_folder = r"D:/para/003"  # 請修改為你的資料夾路徑
    process_folders(base_folder)

if __name__ == "__main__":
    main()