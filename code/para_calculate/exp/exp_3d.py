import numpy as np
import pandas as pd
from scipy.signal import find_peaks
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import warnings
import scipy.signal
from scipy import signal
warnings.filterwarnings('ignore')

class InitialFinalPeakSelector:
    def __init__(self, data):
        self.data = data
        self.initial_peak = None
        self.final_peak = None
        self.fig = None
        self.ax = None
        self.stage = 'initial'  # 'initial' or 'final'
        
    def onclick(self, event):
        if event.button == 1 and event.inaxes:  # Left click
            x = int(round(event.xdata))
            if x >= 0 and x < len(self.data):
                if self.stage == 'initial':
                    self.initial_peak = (x, self.data[x])
                    self.ax.plot(x, self.data[x], 'go', markersize=10)
                    self.ax.text(x, self.data[x], 'Initial Peak', fontsize=10)
                    self.stage = 'final'
                    self.ax.set_title('Now select the final peak (Left click)')
                elif self.stage == 'final':
                    self.final_peak = (x, self.data[x])
                    self.ax.plot(x, self.data[x], 'ro', markersize=10)
                    self.ax.text(x, self.data[x], 'Final Peak', fontsize=10)
                    plt.close()
                self.fig.canvas.draw()
    
    def select_peaks(self):
        self.fig, self.ax = plt.subplots(figsize=(28, 15))
        self.ax.plot(self.data, 'b-')
        self.ax.set_title('Select the initial peak (Left click)')
        self.fig.canvas.mpl_connect('button_press_event', self.onclick)
        plt.show()
        
        if self.initial_peak is None or self.final_peak is None:
            return None, None
        
        return (self.initial_peak[0], self.final_peak[0]), (self.initial_peak[1], self.final_peak[1])

class IntermediatePeakSelector:
    def __init__(self, data):
        self.data = data
        self.selected_peaks = []
        self.peak_values = []
        self.fig = None
        self.ax = None
        self.finished = False
        
    def onclick(self, event):
        if event.button == 1 and event.inaxes:  # Left click
            x = int(round(event.xdata))
            if x >= 0 and x < len(self.data):
                self.selected_peaks.append(x)
                self.peak_values.append(self.data[x])
                self.ax.plot(x, self.data[x], 'go', markersize=10)
                self.fig.canvas.draw()
        elif event.button == 3:  # Right click to finish
            self.finished = True
            plt.close()
    
    def select_peaks(self):
        self.fig, self.ax = plt.subplots(figsize=(28, 15))
        self.ax.plot(self.data, 'b-')
        self.ax.set_title('Click to select intermediate peaks (Right click to finish)')
        self.fig.canvas.mpl_connect('button_press_event', self.onclick)
        plt.show()
        
        return np.array(self.selected_peaks), np.array(self.peak_values)

def read_and_process_data(file_path):
    """
    讀取和處理 xlsx 檔案，並返回角度數據
    """
    df = pd.read_excel(file_path)
    angles = pd.to_numeric(df['Processed_Output'], errors='coerce').values
    angles = angles[~np.isnan(angles)]

    # angles = scipy.signal.savgol_filter(angles, 21, 3, mode='nearest')

    cutoff_freq=0.05
    b, a = signal.butter(4, cutoff_freq, 'low')
    angles = signal.filtfilt(b, a, angles)

    angles = scipy.signal.savgol_filter(angles, 57, 3, mode='nearest')
    return angles

def filter_decreasing_peaks(peaks, peak_values):
    """
    篩選出符合遞減趨勢的峰值
    """
    if len(peaks) <= 1:
        return peaks, peak_values
    
    filtered_peaks = [peaks[0]]
    filtered_values = [peak_values[0]]
    current_max = peak_values[0]
    
    for i in range(1, len(peaks)):
        if peak_values[i] < current_max:
            filtered_peaks.append(peaks[i])
            filtered_values.append(peak_values[i])
            current_max = peak_values[i]
    
    return np.array(filtered_peaks), np.array(filtered_values)

def find_intermediate_peaks(data, start_idx, end_idx, prominence=1, distance=10, height=None, width=None):
    """
    在起始點和終點之間尋找峰值
    """
    segment = data[start_idx:end_idx]
    peaks, properties = find_peaks(segment, 
                                 prominence=prominence,
                                 distance=distance,
                                 height=height,
                                 width=width)
    
    # 調整峰值索引以匹配原始數據
    peaks = peaks + start_idx
    peak_values = data[peaks]
    
    # 篩選符合遞減趨勢的峰值
    filtered_peaks, filtered_values = filter_decreasing_peaks(peaks, peak_values)
    
    return filtered_peaks, filtered_values

def exp_func(x, a, b, c):
    """
    指數函數
    """
    return a * np.exp(-b * (x - np.min(x))) + c

def fit_exponential(peaks, peak_values):
    """
    執行指數擬和
    """
    if len(peaks) < 2:
        print("警告：峰值點太少，無法進行可靠的擬和")
        return None, None

    p0 = [
        max(peak_values) - min(peak_values),
        1.0 / (max(peaks) - min(peaks)),
        min(peak_values)
    ]
    
    bounds = (
        [0, 0, min(peak_values) * 0.5],
        [max(peak_values) * 2, 1.0, max(peak_values)]
    )
    
    try:
        popt, pcov = curve_fit(
            exp_func, 
            peaks, 
            peak_values, 
            p0=p0,
            bounds=bounds,
            maxfev=10000
        )
        return popt, pcov
    except Exception as e:
        print(f"擬和錯誤：{str(e)}")
        return None, None

# def plot_exponential_fit(data, peaks, peak_values, folder_path):
#     """
#     繪製指數擬和結果
#     """
#     popt, pcov = fit_exponential(peaks, peak_values)
    
#     if popt is not None:
#         # 創建新的圖表
#         fig = plt.figure(figsize=(20, 12))
#         plt.plot(data, 'b-', label='Original Data', alpha=0.5)
#         plt.plot(peaks, peak_values, 'go', label='Used Peaks', alpha=0.7)
        
#         x_fit = np.linspace(min(peaks), max(peaks), 1000)
#         y_fit = exp_func(x_fit, *popt)
#         plt.plot(x_fit, y_fit, 'r--', label='Exponential Fit')
        
#         plt.xlabel('Time Step')
#         plt.ylabel('Angle (degrees)')
#         plt.title('Exponential Fit to Decreasing Peak Values')
#         plt.legend()
#         plt.grid(True)
        
#         y_min = min(min(data), min(y_fit)) * 0.9
#         y_max = max(max(data), max(y_fit)) * 1.1
#         plt.ylim(y_min, y_max)
        
#         # 先儲存圖片
#         save_path = os.path.join(folder_path, 'exponential_fit.png')
#         plt.savefig(save_path, bbox_inches='tight', dpi=300)
        
#         # 再顯示圖片
#         plt.show()
        
#         # 最後關閉圖表
#         plt.close(fig)
        
#         return popt
#     return None

def export_to_excel(folder_name, data_dict, excel_path='exponential_results_3d.xlsx'):
    """
    將分析結果輸出到Excel檔案
    """
    row_order = [
        'Initial amplitude (a)',
        'Decay rate (b)',
        'Offset (c)',
        'Number of peaks',
        'Peak positions',
        'Peak values'
    ]
    
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=[''] + [folder_name])
        df.loc[:, ''] = row_order
        df.to_excel(excel_path, index=False)
    
    book = load_workbook(excel_path)
    sheet = book.active
    
    next_col = sheet.max_column + 1
    
    sheet.cell(row=1, column=next_col, value=folder_name)
    
    for i, row_name in enumerate(row_order, start=2):
        value = data_dict.get(row_name, '')
        if isinstance(value, (list, np.ndarray)):
            value = ', '.join(map(str, value))
        sheet.cell(row=i, column=next_col, value=value)
    
    book.save(excel_path)

def plot_exponential_fit(data, all_peaks, all_values, folder_path):
    """
    繪製指數擬和結果，只顯示用於擬和的峰值
    """
    # 排除第一個峰值進行擬和
    peaks_for_fit = all_peaks[1:]
    values_for_fit = all_values[1:]
    
    popt, pcov = fit_exponential(peaks_for_fit, values_for_fit)
    
    if popt is not None:
        # 創建新的圖表
        fig = plt.figure(figsize=(20, 12))
        plt.plot(data, 'b-', label='Original Data', alpha=0.5)
        
        # 只繪製用於擬和的峰值點
        plt.plot(peaks_for_fit, values_for_fit, 'go', label='Peaks Used for Fitting', alpha=0.7)
        
        # 繪製擬和曲線
        x_fit = np.linspace(min(peaks_for_fit), max(peaks_for_fit), 1000)
        y_fit = exp_func(x_fit, *popt)
        plt.plot(x_fit, y_fit, 'r--', label='Exponential Fit')
        
        plt.xlabel('Time Step')
        plt.ylabel('Angle (degrees)')
        plt.title('Exponential Fit to Peak Values')
        plt.legend()
        plt.grid(True)
        
        y_min = min(min(data), min(y_fit)) * 0.9
        y_max = max(max(data), max(y_fit)) * 1.1
        plt.ylim(y_min, y_max)
        
        # 為用於擬和的峰值添加標註
        for i, (peak, value) in enumerate(zip(peaks_for_fit, values_for_fit), start=1):
            plt.annotate(f'Peak {i}\ny={value:.2f}', 
                        xy=(peak, value),
                        xytext=(10, 10),
                        textcoords='offset points')
        
        plt.savefig(os.path.join(folder_path, 'exponential_fit_3d.png'))
        plt.show()
        plt.close(fig)
        
        return popt
    return None

def analyze_file(file_path, folder_path, prominence=1, distance=10, height=None, width=None):
    """
    主要分析函數
    """
    try:
        data = read_and_process_data(file_path)
        
        if len(data) == 0:
            print("警告：沒有有效的數據")
            return None

        # 讓使用者選擇起始和結束峰值
        initial_final_selector = InitialFinalPeakSelector(data)
        (start_idx, end_idx), (start_val, end_val) = initial_final_selector.select_peaks()
        
        if start_idx is None or end_idx is None:
            print("未選擇起始或結束峰值")
            return None

        # 在起始和結束點之間尋找中間的峰值
        intermediate_peaks, intermediate_values = find_intermediate_peaks(
            data, start_idx, end_idx,
            prominence=prominence,
            distance=distance,
            height=height,
            width=width
        )

        # 顯示檢測到的峰值
        fig = plt.figure(figsize=(20, 12))
        plt.plot(data, 'b-', label='Original Data', alpha=0.7)
        plt.plot([start_idx, end_idx], [start_val, end_val], 'ro', 
                label='User Selected Peaks', markersize=10)
        if len(intermediate_peaks) > 0:
            plt.plot(intermediate_peaks, intermediate_values, 'go',
                    label='Detected Intermediate Peaks', alpha=0.7)
        plt.title('Detected Peaks (Accept or manually select intermediate peaks)')
        plt.legend()
        plt.grid(True)
        plt.show()
        plt.close(fig)

        # 詢問使用者是否接受自動檢測的峰值
        if len(intermediate_peaks) > 0:
            if not messagebox.askyesno("確認", "是否接受自動檢測到的中間峰值？"):
                selector = IntermediatePeakSelector(data)
                intermediate_peaks, intermediate_values = selector.select_peaks()
        else:
            if messagebox.askyesno("確認", "未檢測到中間峰值，是否要手動選擇？"):
                selector = IntermediatePeakSelector(data)
                intermediate_peaks, intermediate_values = selector.select_peaks()

        # 組合所有峰值
        all_peaks = np.concatenate(([start_idx], intermediate_peaks, [end_idx]))
        all_values = np.concatenate(([start_val], intermediate_values, [end_val]))

        # 進行指數擬和（不包含第一個峰值）
        fit_params = plot_exponential_fit(data, all_peaks, all_values, folder_path)
        
        if fit_params is not None:
            results = {
                'Initial amplitude (a)': fit_params[0],
                'Decay rate (b)': fit_params[1],
                'Offset (c)': fit_params[2],
                'Number of peaks': len(all_peaks),
                'Peak positions': all_peaks,
                'Peak values': all_values
            }
            return results
        
    except Exception as e:
        print(f"分析過程發生錯誤：{str(e)}")
    return None

def process_folders(base_folder):
    """
    處理所有子資料夾
    """
    root = tk.Tk()
    root.withdraw()  # 隱藏主窗口
    
    for folder_name in os.listdir(base_folder):
        folder_path = os.path.join(base_folder, folder_name)
        print(f"正在處理資料夾: {folder_path}")
        
        for fname in os.listdir(folder_path):
            fname_path = os.path.join(folder_path, fname)
            elec_file = os.path.join(fname_path, 'ae_half.xlsx')
            
            if os.path.isdir(fname_path) and os.path.exists(elec_file):
                try:
                    print(f"正在分析: {fname}")
                    results = analyze_file(elec_file, fname_path)
                    
                    if results is not None:
                        if messagebox.askyesno("確認", "是否要將此結果保存到Excel？"):
                            export_to_excel(fname, results)
                            print(f"成功處理並保存 {fname}")
                        else:
                            print(f"已跳過保存 {fname} 的結果")
                    
                except Exception as e:
                    print(f"處理 {fname} 時發生錯誤: {str(e)}")
    
    root.destroy()

def main():
    base_folder = r"D:/para/020"  # 請修改為你的資料夾路徑
    process_folders(base_folder)

if __name__ == "__main__":
    main()